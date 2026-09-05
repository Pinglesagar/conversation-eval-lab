"""The advisory corpus as a spreadsheet: read it, write it, run from it.

WHY THIS EXISTS
---------------
Scenarios are written by the people who know the domain — compliance officers,
sales-training leads, the person who actually knows what a suitability report has
to say. They have a spreadsheet open all day. They do not have a YAML editor, and
asking them to learn indentation rules before they can add a test case is how a
corpus stops growing.

So this module makes a workbook a first-class way in. Four sheets, one row per
thing, and a Reference sheet that spells out every legal value so Excel's own
validation dropdowns can be pointed at it.

THE RULE THAT KEEPS IT HONEST
-----------------------------
An Excel row does not get its own validation. It is converted into **exactly the
mapping the YAML parser produces**, and then handed to the same
`roleplay.corpus.Scenario` model. There is one rulebook, in one place, and this
module cannot drift from it because it does not contain a copy.

`tests/test_roleplay_excel.py` proves it the only way worth proving: it exports
all 70 committed scenarios to a workbook, reads them back, and asserts the models
are equal field for field. A field this module forgot to carry fails that test.

WHICH FORMAT IS THE SOURCE OF TRUTH
-----------------------------------
YAML is, and deliberately. A `.xlsx` is a zip of XML: it does not diff in a pull
request, it cannot be reviewed line by line, and two people editing one cannot
merge. So the intended loop is:

    export  ->  the domain expert edits the workbook  ->  import  ->  review the
    YAML diff  ->  commit

`--to-yaml` writes files that are byte-stable for unchanged rows, so the diff a
reviewer reads is exactly what the expert changed and nothing else.

Running straight from the workbook is supported too — `load_corpus_from_excel` —
because during a working session with a domain expert, converting on every edit
is friction nobody needs. It is a convenience, not the committed path.

THE DEPENDENCY
--------------
`openpyxl`, in the `[excel]` extra rather than the core dependencies: nothing else
in the repository imports it, and a clean checkout still runs the whole suite
without it. It is in `[dev]` as well, so the round-trip test above always runs and
is never quietly skipped.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Iterable, Sequence

from roleplay.corpus import (
    ARG_OPS,
    CORPUS_ROOT,
    MATCH_MODES,
    SUITES,
    TAG_VOCABULARY,
    VALUE_FREE_OPS,
    Corpus,
    CustomerProfile,
    Scenario,
    load_corpus,
    load_profiles,
)

__all__ = [
    "SHEET_SCENARIOS",
    "SHEET_TURNS",
    "SHEET_ASSERTIONS",
    "SHEET_REFERENCE",
    "ASSERTION_KINDS",
    "export_workbook",
    "read_scenarios",
    "load_corpus_from_excel",
    "write_yaml",
    "main",
]

SHEET_SCENARIOS = "Scenarios"
SHEET_TURNS = "Turns"
SHEET_ASSERTIONS = "Assertions"
SHEET_REFERENCE = "Reference"

#: One row per scenario. Order is the column order in the sheet, and it is the
#: reading order a domain expert wants: who and what first, the machinery last.
SCENARIO_COLUMNS: tuple[str, ...] = (
    "id",
    "suite",
    "title",
    "customer",
    "tags",
    "role",
    "human_verdict",
    "reason",
    "jurisdiction",
    "language",
    "regime",
    "kpis",
    "expected_failure_contracts",
    "expected_failure_since",
    "expected_failure_expectation",
    "score_claims",
    "feedback_grounded",
    "no_progress",
    "phrases_regex",
    "notes",
    "extra_json",
)

TURN_COLUMNS: tuple[str, ...] = ("scenario_id", "turn", "text")

ASSERTION_COLUMNS: tuple[str, ...] = (
    "scenario_id",
    "kind",
    "tool",
    "arg",
    "op",
    "value",
    "match",
    "quantifier",
    "note",
)

#: Every assertion kind, and which columns it reads. The Reference sheet prints
#: this table, so the workbook documents itself.
ASSERTION_KINDS: dict[str, str] = {
    "tool_expected": "tool — this tool must be called at least once",
    "tool_forbidden": "tool — this tool must never be called",
    "min_calls": "tool, value (a whole number) — called at least this many times",
    "max_calls": "tool, value (a whole number) — called at most this many times",
    "ordering": "tool (the first one), arg (the one that must follow), value TRUE for strictly adjacent",
    "arg": "tool, arg, op, value, match, quantifier, note — an assertion about one argument",
    "phrase_required": "value — the adviser must say this",
    "phrase_forbidden": "value — the adviser must not say this",
}

_LIST_SEPARATOR = ", "


def _require_openpyxl():  # noqa: ANN202 - a module, deliberately untyped
    try:
        import openpyxl  # noqa: PLC0415
    except ModuleNotFoundError as exc:  # pragma: no cover - exercised by hand
        raise ModuleNotFoundError(
            "the Excel corpus needs openpyxl, which is not a core dependency of this "
            'repository. Install it with: pip install -e ".[excel]"'
        ) from exc
    return openpyxl


# --------------------------------------------------------------------------- #
# Cell helpers. Excel gives back None for an empty cell and native types for the
# rest; every conversion below is explicit so a stray float never reaches a model.
# --------------------------------------------------------------------------- #


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _list(value: Any) -> tuple[str, ...]:
    """A comma-separated cell to a tuple. Empty cell means empty tuple."""
    raw = _text(value)
    if not raw:
        return ()
    return tuple(part.strip() for part in raw.split(",") if part.strip())


def _bool(value: Any, *, default: bool) -> bool:
    if value is None or (isinstance(value, str) and not value.strip()):
        return default
    if isinstance(value, bool):
        return value
    return _text(value).lower() in {"true", "yes", "y", "1"}


def _int(value: Any, *, where: str) -> int:
    if isinstance(value, bool) or value is None:
        raise ValueError(f"{where}: expected a whole number, got {value!r}")
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = _text(value)
    try:
        return int(text)
    except ValueError:
        raise ValueError(f"{where}: expected a whole number, got {value!r}") from None


def _scalar(value: Any) -> Any:
    """An argument's expected value, keeping the type Excel gave it.

    A float that is a whole number becomes an int, because `total: 18.0` and
    `total: 18` compare differently against a tool argument and the spreadsheet
    has no way to express the difference the author meant.
    """
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return value.strip()
    return value


# --------------------------------------------------------------------------- #
# Scenario -> rows
# --------------------------------------------------------------------------- #


def _scenario_row(scenario: Scenario) -> dict[str, Any]:
    ef = scenario.expected_failure
    #: Anything with no column of its own travels as JSON rather than being
    #: dropped. Two committed scenarios use `consistency`; giving it six columns
    #: for two rows would make the sheet worse for the other sixty-eight.
    extra: dict[str, Any] = {}
    if scenario.consistency is not None:
        extra["consistency"] = scenario.consistency.model_dump(exclude_defaults=True)
    if scenario.divergence is not None:
        extra["divergence"] = scenario.divergence.model_dump(exclude_defaults=True)

    return {
        "id": scenario.id,
        "suite": scenario.suite,
        "title": scenario.title,
        "customer": scenario.customer,
        "tags": _LIST_SEPARATOR.join(scenario.tags),
        "role": scenario.trainee.role,
        "human_verdict": scenario.expectation.human_verdict,
        "reason": scenario.expectation.reason,
        "jurisdiction": scenario.jurisdiction or "",
        "language": scenario.language or "",
        "regime": scenario.regime or "",
        "kpis": _LIST_SEPARATOR.join(scenario.kpis),
        "expected_failure_contracts": _LIST_SEPARATOR.join(ef.contracts) if ef else "",
        "expected_failure_since": ef.since if ef else "",
        "expected_failure_expectation": ef.expectation if ef else "",
        "score_claims": scenario.score_claims,
        "feedback_grounded": scenario.feedback_grounded,
        "no_progress": scenario.no_progress,
        "phrases_regex": scenario.trainee_phrases.regex,
        "notes": scenario.notes,
        "extra_json": json.dumps(extra, sort_keys=True) if extra else "",
    }


def _turn_rows(scenario: Scenario) -> list[dict[str, Any]]:
    return [
        {"scenario_id": scenario.id, "turn": i, "text": text}
        for i, text in enumerate(scenario.trainee.turns, start=1)
    ]


def _assertion_rows(scenario: Scenario) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    def row(kind: str, **cells: Any) -> None:
        base = {c: "" for c in ASSERTION_COLUMNS}
        base.update({"scenario_id": scenario.id, "kind": kind}, **cells)
        rows.append(base)

    for tool in scenario.tools.expected:
        row("tool_expected", tool=tool)
    for tool in scenario.tools.forbidden:
        row("tool_forbidden", tool=tool)
    for tool, n in scenario.tools.min_calls.items():
        row("min_calls", tool=tool, value=n)
    for tool, n in scenario.tools.max_calls.items():
        row("max_calls", tool=tool, value=n)
    for order in scenario.tools.ordering:
        row("ordering", tool=order.first, arg=order.then, value=order.strict)
    for spec in scenario.tools.args:
        row(
            "arg",
            tool=spec.tool,
            arg=spec.arg,
            op=spec.op,
            value="" if spec.value is None else spec.value,
            match=spec.match,
            quantifier=spec.quantifier,
            note=spec.label or "",
        )
    for phrase in scenario.trainee_phrases.required:
        row("phrase_required", value=phrase)
    for phrase in scenario.trainee_phrases.forbidden:
        row("phrase_forbidden", value=phrase)
    return rows


# --------------------------------------------------------------------------- #
# rows -> the mapping the YAML parser would have produced
# --------------------------------------------------------------------------- #


def _mapping_from_rows(
    row: dict[str, Any],
    turns: Sequence[str],
    assertions: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    """Build the dict `Scenario(**data)` expects. No validation happens here.

    Every key is omitted when it is empty rather than passed as `None` or `""`,
    so the model's own defaults apply exactly as they do for a YAML file that
    left the block out.
    """
    sid = _text(row.get("id"))
    where = f"scenario {sid or '(no id)'}"

    data: dict[str, Any] = {
        "id": sid,
        "title": _text(row.get("title")),
        "customer": _text(row.get("customer")),
        "trainee": {"turns": tuple(turns)},
        "expectation": {
            "human_verdict": _text(row.get("human_verdict")),
            "reason": _text(row.get("reason")),
        },
    }
    if role := _text(row.get("role")):
        data["trainee"]["role"] = role
    if tags := _list(row.get("tags")):
        data["tags"] = tags
    for key in ("jurisdiction", "language", "regime"):
        if value := _text(row.get(key)):
            data[key] = value
    if kpis := _list(row.get("kpis")):
        data["kpis"] = kpis
    if notes := _text(row.get("notes")):
        data["notes"] = notes

    ef_expectation = _text(row.get("expected_failure_expectation"))
    if ef_expectation:
        ef: dict[str, Any] = {"expectation": ef_expectation}
        if contracts := _list(row.get("expected_failure_contracts")):
            ef["contracts"] = contracts
        if since := _text(row.get("expected_failure_since")):
            ef["since"] = since
        data["expected_failure"] = ef

    if not _bool(row.get("score_claims"), default=True):
        data["score_claims"] = False
    if not _bool(row.get("feedback_grounded"), default=True):
        data["feedback_grounded"] = False
    if _bool(row.get("no_progress"), default=False):
        data["no_progress"] = True

    if extra_json := _text(row.get("extra_json")):
        try:
            extra = json.loads(extra_json)
        except json.JSONDecodeError as exc:
            raise ValueError(f"{where}: extra_json is not valid JSON: {exc}") from None
        if not isinstance(extra, dict):
            raise ValueError(f"{where}: extra_json must be a JSON object")
        data.update(extra)

    tools: dict[str, Any] = {}
    expected: list[str] = []
    forbidden: list[str] = []
    min_calls: dict[str, int] = {}
    max_calls: dict[str, int] = {}
    ordering: list[dict[str, Any]] = []
    args: list[dict[str, Any]] = []
    required_phrases: list[str] = []
    forbidden_phrases: list[str] = []

    for a in assertions:
        kind = _text(a.get("kind"))
        seat = f"{where}, assertion {kind or '(blank)'}"
        if kind == "tool_expected":
            expected.append(_text(a.get("tool")))
        elif kind == "tool_forbidden":
            forbidden.append(_text(a.get("tool")))
        elif kind == "min_calls":
            min_calls[_text(a.get("tool"))] = _int(a.get("value"), where=seat)
        elif kind == "max_calls":
            max_calls[_text(a.get("tool"))] = _int(a.get("value"), where=seat)
        elif kind == "ordering":
            ordering.append(
                {
                    "first": _text(a.get("tool")),
                    "then": _text(a.get("arg")),
                    "strict": _bool(a.get("value"), default=False),
                }
            )
        elif kind == "arg":
            spec: dict[str, Any] = {"tool": _text(a.get("tool")), "arg": _text(a.get("arg"))}
            if op := _text(a.get("op")):
                spec["op"] = op
            if spec.get("op", "eq") not in VALUE_FREE_OPS:
                spec["value"] = _scalar(a.get("value"))
            if match := _text(a.get("match")):
                spec["match"] = match
            if quantifier := _text(a.get("quantifier")):
                spec["quantifier"] = quantifier
            if label := _text(a.get("note")):
                spec["label"] = label
            args.append(spec)
        elif kind == "phrase_required":
            required_phrases.append(_text(a.get("value")))
        elif kind == "phrase_forbidden":
            forbidden_phrases.append(_text(a.get("value")))
        elif kind:
            raise ValueError(
                f"{seat}: unknown assertion kind {kind!r}; legal kinds are "
                f"{sorted(ASSERTION_KINDS)}"
            )

    if expected:
        tools["expected"] = tuple(expected)
    if forbidden:
        tools["forbidden"] = tuple(forbidden)
    if min_calls:
        tools["min_calls"] = min_calls
    if max_calls:
        tools["max_calls"] = max_calls
    if ordering:
        tools["ordering"] = tuple(ordering)
    if args:
        tools["args"] = tuple(args)
    if tools:
        data["tools"] = tools

    phrases: dict[str, Any] = {}
    if required_phrases:
        phrases["required"] = tuple(required_phrases)
    if forbidden_phrases:
        phrases["forbidden"] = tuple(forbidden_phrases)
    if _bool(row.get("phrases_regex"), default=False):
        phrases["regex"] = True
    if phrases:
        data["trainee_phrases"] = phrases

    return data


# --------------------------------------------------------------------------- #
# Writing a workbook
# --------------------------------------------------------------------------- #

_HEADER_FILL = "FF10131A"
_ACCENT = "FF5FCBBB"


def _style_header(ws, columns: Sequence[str], openpyxl) -> None:  # noqa: ANN001
    from openpyxl.styles import Alignment, Font, PatternFill  # noqa: PLC0415

    ws.append(list(columns))
    for cell in ws[1]:
        cell.font = Font(bold=True, color=_ACCENT, name="Menlo", size=10)
        cell.fill = PatternFill("solid", start_color=_HEADER_FILL)
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _fit(ws, widths: dict[str, int]) -> None:  # noqa: ANN001
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def export_workbook(
    path: str | Path,
    *,
    root: Path | str = CORPUS_ROOT,
    corpus: Corpus | None = None,
) -> Path:
    """Write the whole advisory corpus to one workbook.

    Reads the committed YAML, so what lands in the sheet is what the validator
    already accepted — an export can never contain a scenario the harness would
    refuse to load.
    """
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Alignment, Font  # noqa: PLC0415

    loaded = corpus if corpus is not None else load_corpus(root)
    scenarios = sorted(loaded.scenarios, key=lambda s: (s.suite, s.id))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(SHEET_SCENARIOS)
    _style_header(ws, SCENARIO_COLUMNS, openpyxl)
    for scenario in scenarios:
        row = _scenario_row(scenario)
        ws.append([row[c] for c in SCENARIO_COLUMNS])
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _fit(ws, {"A": 42, "B": 12, "C": 46, "D": 20, "E": 30, "F": 26, "G": 14, "H": 60,
              "I": 14, "J": 10, "K": 10, "L": 14, "M": 30, "N": 14, "O": 46,
              "P": 13, "Q": 18, "R": 12, "S": 13, "T": 50, "U": 24})

    ws = wb.create_sheet(SHEET_TURNS)
    _style_header(ws, TURN_COLUMNS, openpyxl)
    for scenario in scenarios:
        for turn in _turn_rows(scenario):
            ws.append([turn[c] for c in TURN_COLUMNS])
    for row_cells in ws.iter_rows(min_row=2):
        row_cells[2].alignment = Alignment(vertical="top", wrap_text=True)
    _fit(ws, {"A": 42, "B": 7, "C": 120})

    ws = wb.create_sheet(SHEET_ASSERTIONS)
    _style_header(ws, ASSERTION_COLUMNS, openpyxl)
    for scenario in scenarios:
        for a in _assertion_rows(scenario):
            ws.append([a[c] for c in ASSERTION_COLUMNS])
    for row_cells in ws.iter_rows(min_row=2):
        row_cells[5].alignment = Alignment(vertical="top", wrap_text=True)
    _fit(ws, {"A": 42, "B": 18, "C": 22, "D": 18, "E": 10, "F": 60, "G": 12, "H": 12, "I": 28})

    _write_reference(wb, loaded.profiles, openpyxl)

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _write_reference(wb, profiles: dict[str, CustomerProfile], openpyxl) -> None:  # noqa: ANN001
    """The sheet that makes the workbook self-documenting.

    Every legal value, in a column, so Excel's own Data Validation can be pointed
    at a range here and turn the authoring sheets into pick-lists.
    """
    from openpyxl.styles import Alignment, Font  # noqa: PLC0415

    ws = wb.create_sheet(SHEET_REFERENCE)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 96

    def head(text: str) -> None:
        ws.append([text, ""])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color=_ACCENT, name="Menlo", size=10)

    def line(a: str, b: str = "") -> None:
        ws.append([a, b])
        ws.cell(ws.max_row, 2).alignment = Alignment(vertical="top", wrap_text=True)

    head("HOW THIS WORKBOOK WORKS")
    line("Scenarios", "One row per test case. The id must start with its suite name and a hyphen.")
    line("Turns", "What the adviser says, one line per turn, in order, linked by scenario_id.")
    line("Assertions", "Everything that must be true afterwards. The 'kind' column decides which other columns are read.")
    line("", "")
    line("Import it back with", "python -m roleplay.excel_corpus import <file.xlsx> --to-yaml")
    line("Check it without writing", "python -m roleplay.excel_corpus check <file.xlsx>")
    line("", "")

    head("ASSERTION KINDS")
    for kind, description in ASSERTION_KINDS.items():
        line(kind, description)
    line("", "")

    head("SUITES")
    for suite in SUITES:
        line(suite, f"ids in this suite must begin '{suite}-'")
    line("", "")

    head("CUSTOMERS")
    for key, profile in sorted(profiles.items()):
        line(key, profile.display_name)
    line("", "")

    head("TAGS")
    for tag, meaning in TAG_VOCABULARY.items():
        line(tag, meaning)
    line("", "")

    head("ARGUMENT OPERATORS")
    for op in sorted(ARG_OPS):
        note = "needs no value" if op in VALUE_FREE_OPS else ""
        line(op, note)
    line("", "")

    head("MATCH MODES")
    for mode in sorted(MATCH_MODES):
        line(mode)
    line("", "")

    head("A WARNING WORTH READING")
    line("Excel changes some values", "A code typed as 007 is stored as 7; 1-2 can become a date; a long account number becomes 1.2E+15. Format such cells as Text before typing, and check the imported YAML diff.")
    line("YAML stays the source of truth", "This workbook is for authoring. What gets committed and reviewed is the YAML the import writes.")


# --------------------------------------------------------------------------- #
# Reading a workbook
# --------------------------------------------------------------------------- #


def _sheet_rows(ws, columns: Sequence[str]) -> list[dict[str, Any]]:  # noqa: ANN001
    header = [(_text(c.value) or "") for c in ws[1]]
    missing = [c for c in columns if c not in header]
    if missing:
        raise ValueError(
            f"sheet {ws.title!r} is missing column(s) {missing}; expected {list(columns)}"
        )
    index = {name: i for i, name in enumerate(header)}
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            continue  # a blank spacer row is not an error
        rows.append({name: values[i] if i < len(values) else None for name, i in index.items()})
    return rows


def read_scenarios(path: str | Path) -> tuple[Scenario, ...]:
    """Read a workbook into validated `Scenario` models.

    Every row goes through `roleplay.corpus.Scenario`, the same model the YAML
    loader uses, so a workbook cannot express anything a YAML file could not — and
    a bad row fails here with the same message it would have failed with there.
    """
    openpyxl = _require_openpyxl()
    src = Path(path)
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    try:
        for name in (SHEET_SCENARIOS, SHEET_TURNS, SHEET_ASSERTIONS):
            if name not in wb.sheetnames:
                raise ValueError(
                    f"{src}: missing sheet {name!r}; the workbook has {wb.sheetnames}"
                )
        scenario_rows = _sheet_rows(wb[SHEET_SCENARIOS], SCENARIO_COLUMNS)
        turn_rows = _sheet_rows(wb[SHEET_TURNS], TURN_COLUMNS)
        assertion_rows = _sheet_rows(wb[SHEET_ASSERTIONS], ASSERTION_COLUMNS)
    finally:
        wb.close()

    turns: dict[str, list[tuple[int, str]]] = {}
    for row in turn_rows:
        sid = _text(row.get("scenario_id"))
        turns.setdefault(sid, []).append(
            (_int(row.get("turn"), where=f"{sid}: turn number"), _text(row.get("text")))
        )
    assertions: dict[str, list[dict[str, Any]]] = {}
    for row in assertion_rows:
        assertions.setdefault(_text(row.get("scenario_id")), []).append(row)

    seen: set[str] = set()
    out: list[Scenario] = []
    for row in scenario_rows:
        sid = _text(row.get("id"))
        if not sid:
            raise ValueError(f"{src}: a Scenarios row has no id")
        if sid in seen:
            raise ValueError(f"{src}: duplicate scenario id {sid!r}")
        seen.add(sid)
        suite = _text(row.get("suite"))
        if suite not in SUITES:
            raise ValueError(f"{sid}: suite {suite!r} is not one of {list(SUITES)}")
        ordered = [text for _, text in sorted(turns.get(sid, []), key=lambda t: t[0])]
        data = _mapping_from_rows(row, ordered, assertions.get(sid, []))
        scenario = Scenario(**data, suite=suite, source=f"{src}#{sid}")
        if not scenario.id.startswith(f"{suite}-"):
            raise ValueError(
                f"id {scenario.id!r} must start with its suite prefix {suite!r}- "
                "(the same rule the YAML loader enforces through the directory name)"
            )
        out.append(scenario)

    unknown = set(turns) | set(assertions) - seen
    stray = sorted(x for x in unknown if x and x not in seen)
    if stray:
        raise ValueError(
            f"{src}: Turns/Assertions rows reference scenario id(s) not in the "
            f"Scenarios sheet: {stray}"
        )
    return tuple(out)


def load_corpus_from_excel(
    path: str | Path, *, profiles_dir: Path | str | None = None
) -> Corpus:
    """A `Corpus` built from a workbook, ready for anything that takes one.

    Customer profiles still come from YAML: they are shared across scenarios,
    rarely edited, and a profile is a nested structure that a flat sheet models
    badly. The workbook references them by key and the check below is the same
    one `load_corpus` makes.
    """
    profiles = load_profiles(Path(profiles_dir) if profiles_dir else CORPUS_ROOT / "customers")
    scenarios = read_scenarios(path)
    for scenario in scenarios:
        if scenario.customer not in profiles:
            raise ValueError(
                f"{scenario.id}: unknown customer profile {scenario.customer!r}; "
                f"have {sorted(profiles)}"
            )
    return Corpus(scenarios=scenarios, profiles=profiles)


# --------------------------------------------------------------------------- #
# Writing YAML back out
# --------------------------------------------------------------------------- #

#: The order keys are written in. It is the order a person reads a scenario in —
#: who and what, then what was said, then what must be true — and it is fixed so
#: that re-importing an unchanged workbook produces an unchanged file and the
#: reviewer's diff shows only what the author actually edited.
_YAML_KEY_ORDER: tuple[str, ...] = (
    "id",
    "title",
    "customer",
    "tags",
    "jurisdiction",
    "language",
    "regime",
    "kpis",
    "trainee",
    "expectation",
    "tools",
    "trainee_phrases",
    "score_claims",
    "feedback_grounded",
    "no_progress",
    "consistency",
    "divergence",
    "expected_failure",
    "notes",
)


def _yaml_ready(scenario: Scenario) -> dict[str, Any]:
    """The scenario as plain data, without the fields the loader supplies itself.

    `suite` comes from the directory and `source` from the path, so writing either
    into the file would create a second place for them to disagree with reality.
    """
    data = scenario.model_dump(exclude_defaults=True, exclude_none=True)
    data.pop("suite", None)
    data.pop("source", None)
    data["id"] = scenario.id
    data["title"] = scenario.title
    data["customer"] = scenario.customer
    data["trainee"] = {"role": scenario.trainee.role, "turns": list(scenario.trainee.turns)}
    data["expectation"] = {
        "human_verdict": scenario.expectation.human_verdict,
        "reason": scenario.expectation.reason,
    }
    return {k: data[k] for k in _YAML_KEY_ORDER if k in data}


def _plain(value: Any) -> Any:
    if isinstance(value, tuple):
        return [_plain(v) for v in value]
    if isinstance(value, list):
        return [_plain(v) for v in value]
    if isinstance(value, dict):
        return {k: _plain(v) for k, v in value.items()}
    return value


def write_yaml(
    path: str | Path, *, root: Path | str = CORPUS_ROOT, dry_run: bool = False
) -> dict[str, str]:
    """Convert a workbook into the YAML corpus. Returns {path: 'written'|'unchanged'|'would write'}.

    A file is rewritten only when the workbook row *means* something different
    from the file already on disk — compared as loaded `Scenario` models, not as
    text. That distinction is the whole reason the review loop works.

    The committed YAML is hand-written: folded block scalars, inline flow maps,
    comments between blocks. A generator cannot reproduce that byte for byte, and
    if it rewrote every file on every import the reviewer would face seventy files
    of formatting churn with the one real edit buried in it. Comparing meaning
    instead means an untouched row leaves its file — and its comments, and its
    `git blame` — exactly as they were, and only an edited scenario appears in the
    diff. That file does get reformatted, which is a visible and acceptable cost
    on a file somebody deliberately changed.
    """
    import yaml  # noqa: PLC0415

    from roleplay.corpus import load_scenario  # noqa: PLC0415

    scenarios = read_scenarios(path)
    base = Path(root)
    result: dict[str, str] = {}
    for scenario in scenarios:
        target = base / scenario.suite / f"{scenario.id}.yaml"
        if target.is_file():
            try:
                current = load_scenario(target)
            except Exception:  # noqa: BLE001 - an unreadable file is one to overwrite
                current = None
            if current is not None and _same_meaning(current, scenario):
                result[str(target)] = "unchanged"
                continue
        if dry_run:
            result[str(target)] = "would write"
            continue
        text = yaml.safe_dump(
            _plain(_yaml_ready(scenario)),
            sort_keys=False,
            allow_unicode=True,
            default_flow_style=False,
            width=88,
        )
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(text, encoding="utf-8")
        result[str(target)] = "written"
    return result


def _same_meaning(a: Scenario, b: Scenario) -> bool:
    """Every field but the two the loader supplies from the file's own location."""
    skip = {"source", "suite"}
    return all(
        getattr(a, name) == getattr(b, name) for name in Scenario.model_fields if name not in skip
    )


# --------------------------------------------------------------------------- #
# The command line
# --------------------------------------------------------------------------- #


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="python -m roleplay.excel_corpus",
        description=(
            "The advisory scenario corpus as a spreadsheet. Export it, hand the "
            "workbook to the person who knows the domain, import their edits back."
        ),
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p_export = sub.add_parser("export", help="write the committed YAML corpus to a workbook")
    p_export.add_argument("path", type=Path, nargs="?", default=Path("roleplay_scenarios.xlsx"))
    p_export.add_argument("--root", type=Path, default=CORPUS_ROOT)

    p_check = sub.add_parser("check", help="validate a workbook and write nothing")
    p_check.add_argument("path", type=Path)

    p_import = sub.add_parser("import", help="convert a workbook into YAML scenario files")
    p_import.add_argument("path", type=Path)
    p_import.add_argument("--to-yaml", action="store_true", help="actually write the files")
    p_import.add_argument("--root", type=Path, default=CORPUS_ROOT)

    args = parser.parse_args(argv)

    if args.command == "export":
        out = export_workbook(args.path, root=args.root)
        size = out.stat().st_size
        print(f"wrote {out}  ({size:,} bytes)")
        print(f"  sheets: {SHEET_SCENARIOS}, {SHEET_TURNS}, {SHEET_ASSERTIONS}, {SHEET_REFERENCE}")
        print("  YAML remains the source of truth; import the workbook back and review the diff.")
        return 0

    if args.command == "check":
        try:
            scenarios = read_scenarios(args.path)
        except Exception as exc:  # noqa: BLE001 - the message is the product here
            print(f"REFUSED  {args.path}\n  {exc}", file=sys.stderr)
            return 1
        by_suite: dict[str, int] = {}
        for s in scenarios:
            by_suite[s.suite] = by_suite.get(s.suite, 0) + 1
        turns = sum(len(s.trainee.turns) for s in scenarios)
        print(f"OK  {args.path}")
        print(f"  {len(scenarios)} scenarios, {turns} adviser turns")
        for suite in SUITES:
            if suite in by_suite:
                print(f"    {suite:12} {by_suite[suite]}")
        return 0

    result = write_yaml(args.path, root=args.root, dry_run=not args.to_yaml)
    changed = {k: v for k, v in result.items() if v != "unchanged"}
    print(f"{len(result)} scenario(s); {len(changed)} would change" if not args.to_yaml
          else f"{len(result)} scenario(s); {len(changed)} written")
    for target, state in sorted(changed.items()):
        print(f"  {state:12} {target}")
    if not args.to_yaml and changed:
        print("\nnothing was written — re-run with --to-yaml to apply, then review the git diff")
    return 0


if __name__ == "__main__":
    sys.exit(main())
